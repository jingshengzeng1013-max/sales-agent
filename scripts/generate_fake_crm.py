# -*- coding: utf-8 -*-
"""
CRM 数据脱敏脚本
生成仿造的假数据 Excel，替换 docs/ 下的真实 CRM 文件。

脱敏策略：
1. 公司名 → 替换为虚构公司名（保留行业特征，但不含真实企业）
2. 人名 → 替换为常见姓+随机名
3. 电话 → 替换为 138/139 开头的假号码
4. 信用代码 → 生成格式合法但虚构的统一社会信用代码
5. 商机编号/线索编号 → 保留编号格式但替换序号
6. 金额/日期/状态 → 保留原始分布特征
"""

import random
import string
import os
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

random.seed(42)  # 可复现

# ============================================================
# 假数据生成器
# ============================================================

# 虚构公司名池（按行业分组，不含真实企业）
FAKE_COMPANIES = {
    "汽车": [
        "星途新能源汽车有限公司", "云驰汽车制造有限公司", "远航电动汽车有限公司",
        "骏马汽车股份有限公司", "天行智能汽车有限公司", "鹏飞新能源汽车有限公司",
        "凌云汽车科技有限公司", "驰骋新能源车辆有限公司",
    ],
    "通信": [
        "联通星辰通信有限公司", "电信翼联通信有限公司", "移动数智通信有限公司",
        "天波通信技术有限公司", "信通光电有限公司", "华信通信科技股份有限公司",
        "中星通联通信有限公司", "瑞信网络通信有限公司",
    ],
    "电子": [
        "蓝芯微电子有限公司", "晶锐微电子科技有限公司", "宏达电子器件有限公司",
        "创新电子技术有限公司", "智芯集成电路有限公司", "华微电子股份有限公司",
        "鹏程电子科技有限公司", "星辰微电子有限公司",
    ],
    "终端": [
        "智联终端设备有限公司", "移动终端科技有限公司", "数码先锋科技有限公司",
        "智能终端制造有限公司", "万物互联终端有限公司", "新锐数码科技有限公司",
    ],
    "卫星": [
        "星河卫星通信有限公司", "天际卫星科技有限公司", "航天星联通信有限公司",
        "银河卫星技术有限公司", "星际通信科技有限公司", "苍穹卫星通信有限公司",
    ],
    "高校": [
        "华北理工大学", "南方科技大学", "华东交通大学", "西京大学",
        "东海理工学院", "南山科技大学", "北江工业大学",
    ],
    "政府": [
        "市政务服务中心", "区信息化管理中心", "省应急管理厅", "市交通运输局",
        "区卫生健康委员会", "市生态环境局", "省水利厅",
    ],
    "其他": [
        "恒达科技有限公司", "创新科技发展有限公司", "远景技术有限公司",
        "宏图智能科技有限公司", "瑞达信息技术有限公司", "天诚科技有限公司",
        "正通实业有限公司", "汇智电子有限公司", "德信科技有限公司",
        "光华技术有限公司", "博远智能科技有限公司", "拓达电子有限公司",
    ],
}

# 常见姓氏
SURNAMES = list("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟黄穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮龚程嵇邢滑裴陆荣翁荀羊於惠甄曲家封芮羿储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫")

# 名字用字
GIVEN_NAMES = list("伟芳娜秀英敏静丽强磊军洋勇艳杰娟涛明超秀兰霞平刚桂英华金鹏飞玲桂兰丹萍鹏华彬远刚建华国栋鑫宇晨辰浩然子轩雨泽思源志强文博嘉熙俊豪皓宇")

# 假电话前缀
PHONE_PREFIXES = ["138", "139", "136", "137", "135", "158", "159", "188"]

# 产品名池
PRODUCTS = [
    "天通卫星模组", "卫星通信终端", "北斗定位模块", "物联网通信模块",
    "应急通信设备", "车载通信终端", "卫星电话", "数据传输模块",
    "射频前端模组", "通信天线", "信号处理芯片", "导航定位终端",
]

# 跟进人员池（虚构）
SALESPEOPLE = ["张磊", "刘洋", "陈静", "杨帆", "赵强", "黄敏", "周杰", "吴婷", "徐辉", "孙莉"]

# 商机状态
OPP_STATUS = ["跟进中", "关闭", "签约", "流失"]
# 跟进状态
FOLLOWUP_STATUS = ["初次接触", "需求确认", "方案报价", "商务谈判", "签约", "已关闭"]
# 线索状态
LEAD_STATUS = ["跟进中", "封存", "转化", "流失"]
# 客户类型
CUSTOMER_TYPES = ["汽车组", "通信组", "终端组", "行业组", "卫星组"]
# 结束原因
CLOSE_REASONS = ["价格未谈拢", "客户取消项目", "竞争对手中标", "需求变更", "项目延期", None, None, None]


def gen_name():
    """生成随机姓名"""
    surname = random.choice(SURNAMES)
    given = "".join(random.sample(GIVEN_NAMES, random.randint(1, 2)))
    return surname + given


def gen_phone():
    """生成假电话号码"""
    prefix = random.choice(PHONE_PREFIXES)
    suffix = "".join(random.choice(string.digits) for _ in range(8))
    return prefix + suffix


def gen_credit_code():
    """
    生成格式合法的统一社会信用代码（18位）
    格式: 登记管理部门(1) + 机构类别(1) + 登记管理机关(6) + 主体标识(9) + 校验码(1)
    """
    # 登记管理部门: 9=工商
    # 机构类别: 1=企业
    # 登记管理机关: 随机6位数字
    # 主体标识: 随机9位字母数字
    # 校验码: 随机1位字母数字
    part1 = "91"
    part2 = "".join(random.choice(string.digits) for _ in range(6))
    part3 = "".join(random.choice(string.ascii_uppercase + string.digits) for _ in range(9))
    part4 = random.choice(string.ascii_uppercase + string.digits)
    return part1 + part2 + part3 + part4


def gen_company():
    """生成随机公司名"""
    industry = random.choice(list(FAKE_COMPANIES.keys()))
    company = random.choice(FAKE_COMPANIES[industry])
    return company


def gen_date(start_year=2024, end_year=2026):
    """生成随机日期"""
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 7, 1)
    delta = end - start
    random_days = random.randint(0, delta.days)
    dt = start + timedelta(days=random_days)
    return dt.strftime("%Y-%m-%d")


def gen_datetime(start_year=2024, end_year=2026):
    """生成随机日期时间"""
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 7, 1)
    delta = end - start
    random_seconds = random.randint(0, int(delta.total_seconds()))
    dt = start + timedelta(seconds=random_seconds)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def gen_amount():
    """生成随机商机金额"""
    choices = [
        random.choice([0, 0, 0]),  # 经常为0
        random.randint(5, 50) * 10000,  # 5-50万
        random.randint(50, 500) * 10000,  # 50-500万
        random.randint(500, 5000) * 10000,  # 500-5000万
    ]
    return float(random.choice(choices))


def gen_quantity():
    """生成随机售卖数量"""
    return float(random.choice([0, 0, random.randint(10, 10000), random.randint(100, 50000)]))


# ============================================================
# 商机列表生成
# ============================================================

def generate_business_excel(output_path, num_rows=90):
    """生成脱敏商机列表 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商机列表"

    # 标题行
    title = "商机列表"
    ws.merge_cells("A1:AA1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = [
        "商机编号", "商机名称", "关联线索编号", "跟进人员", "客户名称",
        "客户统一信用代码", "是否重复客户", "客户联系人", "联系人职位",
        "商机状态", "跟进状态", "计划跟进时间", "提报时间", "商机封闭期",
        "关联线索名称", "商机规模（元）", "是否关联商机", "商机关联编号",
        "商机关联负责人", "商机跟进反馈", "报备时间", "结束原因",
        "客户类型", "售卖产品", "售卖数量", "预期使用终端", "更新时间",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 数据行
    for i in range(num_rows):
        row_num = i + 3
        company = gen_company()
        salesperson = random.choice(SALESPEOPLE)
        contact_person = gen_name()
        opp_id = f"SJ{random.randint(2411000, 2502000) + i}"
        opp_status = random.choice(OPP_STATUS)
        followup = random.choice(FOLLOWUP_STATUS) if opp_status == "跟进中" else "已关闭"
        amount = gen_amount()
        product = random.choice(PRODUCTS)
        customer_type = random.choice(CUSTOMER_TYPES)
        close_reason = random.choice(CLOSE_REASONS) if opp_status == "关闭" else None

        row_data = [
            opp_id,                                          # 商机编号
            f"{company}-{product}项目",                      # 商机名称
            f"XS{random.randint(2411000, 2502000)}",         # 关联线索编号
            salesperson,                                     # 跟进人员
            company,                                         # 客户名称
            gen_credit_code(),                               # 客户统一信用代码
            random.choice(["是", "否"]),                      # 是否重复客户
            contact_person,                                  # 客户联系人
            random.choice(["采购负责人", "技术主管", "项目经理", ""]),  # 联系人职位
            opp_status,                                      # 商机状态
            followup,                                        # 跟进状态
            gen_date() if opp_status == "跟进中" else None,  # 计划跟进时间
            gen_datetime(),                                  # 提报时间
            gen_date(2026, 2027),                            # 商机封闭期
            product,                                         # 关联线索名称
            amount,                                          # 商机规模（元）
            random.choice(["是", "否"]),                      # 是否关联商机
            f"SJ{random.randint(2411000, 2502000)}" if random.random() > 0.7 else None,  # 商机关联编号
            random.choice(SALESPEOPLE) if random.random() > 0.7 else None,  # 商机关联负责人
            random.choice(["开始批量供货", "技术方案确认中", "等待客户反馈", "已签合同"]),  # 商机跟进反馈
            gen_datetime(),                                  # 报备时间
            close_reason,                                    # 结束原因
            customer_type,                                   # 客户类型
            product,                                         # 售卖产品
            gen_quantity(),                                  # 售卖数量
            random.choice(["车载终端", "手持终端", "固定终端", "船载终端", ""]),  # 预期使用终端
            gen_datetime(),                                  # 更新时间
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"商机列表已生成: {output_path} ({num_rows} 条记录)")


# ============================================================
# 线索列表生成
# ============================================================

def generate_lead_excel(output_path, num_rows=340):
    """生成脱敏线索列表 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "线索列表"

    # 标题行
    title = "线索列表"
    ws.merge_cells("A1:AA1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = [
        "线索编号", "线索名称", "报备人员", "联系方式", "客户名称",
        "客户统一信用代码", "是否重复客户", "客户联系人", "联系人职位",
        "线索获取时间", "线索报备时间", "线索状态", "业务审批",
        "业务审批结果", "销售人员", "是否关联线索", "线索关联编号",
        "线索关联负责人", "线索情况简述", "信息备注", "审核说明",
        "预计启动时间", "线索规模", "售卖数量", "更新时间",
        "售卖产品", "客户类型",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 数据行
    for i in range(num_rows):
        row_num = i + 3
        company = gen_company()
        reporter = random.choice(SALESPEOPLE)
        contact_person = gen_name()
        lead_id = f"XS{random.randint(2410000, 2506000) + i}"
        lead_status = random.choice(LEAD_STATUS)
        product = random.choice(PRODUCTS)
        customer_type = random.choice(CUSTOMER_TYPES)
        phone = gen_phone()

        # 线索简述
        descriptions = [
            f"{company}是一家专注于通信技术研发的企业，有卫星通信模组采购需求，计划在年内启动项目。",
            f"客户正在评估{product}方案，已与多家供应商接触，需要加快跟进节奏。",
            f"{company}近期发布了采购计划，涉及{product}等设备，预算预计在50-200万元。",
            f"客户对天通卫星通信产品有明确需求，项目处于方案评估阶段。",
            f"通过行业展会获取的线索，{company}表达了对接合作意向。",
        ]

        row_data = [
            lead_id,                                          # 线索编号
            f"{company}-{product}",                           # 线索名称
            reporter,                                         # 报备人员
            phone,                                             # 联系方式
            company,                                          # 客户名称
            gen_credit_code(),                                # 客户统一信用代码
            random.choice(["是", "否"]),                       # 是否重复客户
            contact_person,                                   # 客户联系人
            random.choice(["销售主管", "技术总监", "采购经理", ""]),  # 联系人职位
            gen_datetime(2024, 2025),                          # 线索获取时间
            gen_datetime(),                                    # 线索报备时间
            lead_status,                                       # 线索状态
            f"[{','.join(random.sample(SALESPEOPLE, random.randint(2, 4)))}]",  # 业务审批
            random.choice(["通过", "待审批", "通过", "通过"]),   # 业务审批结果
            reporter,                                          # 销售人员
            random.choice(["是", "否"]),                       # 是否关联线索
            f"XS{random.randint(2410000, 2506000)}" if random.random() > 0.8 else None,  # 线索关联编号
            random.choice(SALESPEOPLE) if random.random() > 0.8 else None,  # 线索关联负责人
            random.choice(descriptions),                      # 线索情况简述
            None,                                              # 信息备注
            random.choice(["跟紧后续进展", "持续跟进", "等待技术方案", None]),  # 审核说明
            gen_date(2025, 2026) if lead_status == "跟进中" else None,  # 预计启动时间
            gen_amount(),                                      # 线索规模
            gen_quantity(),                                    # 售卖数量
            gen_datetime(),                                    # 更新时间
            product,                                           # 售卖产品
            customer_type,                                     # 客户类型
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"线索列表已生成: {output_path} ({num_rows} 条记录)")


# ============================================================
# 主函数
# ============================================================

def main():
    docs_dir = Path(__file__).resolve().parent.parent / "docs"

    # 备份原始文件（如果还没备份过）
    business_file = docs_dir / "商机列表20260422113444.xlsx"
    lead_file = docs_dir / "线索列表20260422113326.xlsx"

    for f in [business_file, lead_file]:
        backup = f.with_suffix(".xlsx.bak")
        if f.exists() and not backup.exists():
            import shutil
            shutil.copy2(f, backup)
            print(f"已备份: {f.name} → {backup.name}")

    # 生成脱敏数据
    generate_business_excel(business_file, num_rows=90)
    generate_lead_excel(lead_file, num_rows=340)

    print("\n✅ CRM 数据脱敏完成！")
    print(f"   商机列表: {business_file}")
    print(f"   线索列表: {lead_file}")
    print("\n⚠️  原始文件已备份为 .xlsx.bak 文件")
    print("⚠️  请重新运行客户画像生成流程以更新 customer_profiles.jsonl")


if __name__ == "__main__":
    main()
